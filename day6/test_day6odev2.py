from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl
from constants import globalConstants as gc
from time import sleep


class Test_DemoClass:
    #her testten önce çağrılır
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(gc.URL)
        # günün tarihini al bu tarih ile bir klasör var mı kontrol et yoksa oluştur
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)
       
    
    #her testten sonra çağrılır
    def teardown_method(self):
        self.driver.quit()
    
    def invalidData():
        excelFile = openpyxl.load_workbook("data/invalid_login.xlsx")
        selectedSheet = excelFile["Sayfa1"]

        totalRows = selectedSheet.max_row
        data=[]
        for i in range(2, totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)
        return data
    
   
    #Kod tekrarını azaltmak için tekrarlayan işlemler fonksiyonlara alındı.Tekrar eden textler ise constants projesine alındı.
    def waitForElementVisible(self,locator,timeout=5):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))
    

    def returnErrorMesagge(self):
        errorMessage=self.driver.find_element(By.XPATH,"//div[@class='error-message-container error']/h3")
        return errorMessage
    
    def returnUserName(self):
        userNameInput=self.driver.find_element(By.ID , "user-name")
        return userNameInput
    
    def returnPaswword(self):
        passwordInput =self.driver.find_element(By.ID , "password")
        return passwordInput
    
    def invalidData():
        excelFile = openpyxl.load_workbook("data/invalid_login.xlsx")
        selectedSheet = excelFile["Sayfa1"]

        totalRows = selectedSheet.max_row
        data=[]
        for i in range(2, totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)
        return data
   
   
  

    @pytest.mark.parametrize("username,password",[("standard_user","secret_sauce")])
    def test_standart_user(self,username,password):
        self.waitForElementVisible((By.ID,"user-name"))
        self.waitForElementVisible((By.ID,"password"),10)
        self.returnUserName().send_keys(username)
        self.returnPaswword().send_keys(password)
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        self.driver.save_screenshot(f"{self.folderPath}/test-standart-user-{username}-{password}.png")
       
  
    def test_numberOfProduct(self):
        self.waitForElementVisible((By.ID,"user-name"))
        self.waitForElementVisible((By.ID,"password"),10)
        self.returnUserName().send_keys("standard_user")
        self.returnPaswword().send_keys("secret_sauce")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        self.waitForElementVisible((By.CLASS_NAME,"inventory_item"))
        number_products =self.driver.find_elements(By.CLASS_NAME, "inventory_item")
        print(f"Standart kullanıcıya gösterilen toplam üyün sayısı: {len(number_products)}")
        self.driver.save_screenshot(f"{self.folderPath}/test_numberOfProduct-standard_user-secret_sauce.png")
        assert True
      
    
    @pytest.mark.parametrize("username,password",invalidData())
    def test_invalid_login(self,username,password):
        self.waitForElementVisible((By.ID,"user-name"))
        self.waitForElementVisible((By.ID,"password"),10)
        self.returnUserName().send_keys(username)
        self.returnPaswword().send_keys(password)
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage=self.returnErrorMesagge()
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")
        assert errorMessage.text == gc.invalidlogin_errmsg
   
 
    @pytest.mark.parametrize("username,password",[("problem_user","secret_sauce")])
    def test_others_user(self,username,password):
        self.waitForElementVisible((By.ID,"user-name"))
        self.waitForElementVisible((By.ID,"password"),10)
        self.returnUserName().send_keys(username)
        self.returnPaswword().send_keys(password)
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        self.driver.save_screenshot(f"{self.folderPath}/test-others-user-{username}-{password}.png")
    
   
    
    @pytest.mark.parametrize("username,password",[("problem_user","secret_sauce")])
    def test_others_user_Shop(self,username,password):
        self.waitForElementVisible((By.ID,"user-name"))
        self.waitForElementVisible((By.ID,"password"),10)
        self.returnUserName().send_keys(username)
        self.returnPaswword().send_keys(password)
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        
        kart_buton=self.driver.find_element(By.XPATH,"/html/body/div/div/div/div[2]/div/div/div/div[2]/div[2]/div[2]/button")
        kart_buton.click()
        self.driver.save_screenshot(f"{self.folderPath}/test-others-user-Shop-{username}-{password}.png")

        
    
        
        
        
        


        
    